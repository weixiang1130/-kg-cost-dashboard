#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
KG 預算發包請款結算一覽表 → 工程造價分析 v14
============================================
核心邏輯：
  1. 結構項目數量用「核心工項篩選」+「發包數量」
  2. 混凝土子項用 Regex 篩選材料本體，其餘歸「其他混凝土工程」
  3. D.10/D.00/D.13 合併為「其他雜項工程」放於陸門窗下
  4. 單價保留 1 位小數
  5. 空 C-code 自動反查（描述欄 / 合約 / 繼承三層 fallback）

此模組同時可 CLI 執行，或被 kg_dashboard.py 匯入使用。
"""
import os, sys, glob, argparse, csv, re, zipfile, io
from collections import defaultdict
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Border, Side
from openpyxl.utils import get_column_letter

PING_PER_M2 = 0.3025

# ══════════════════════════════════════════
# 核心工項篩選規則 — 用於提取代表性數量
# ══════════════════════════════════════════
QTY_RULES = {
    'C.02.00.00': {'unit': 'M2', 'core': ['模板','免拆'], 'excl': ['放樣','支撐','壓送','點工','獎金','補貼','遮斷','買斷','自走車','角材']},
    'C.03.01.00': {'unit': 'T|噸', 'core': ['鋼筋'], 'excl': ['獎金','補貼','運費']},
    'C.03.02.00': {'unit': 'T|噸', 'core': ['加工','鋼筋'], 'excl': ['獎金','補貼']},
    'C.03.03.00': {'unit': 'T|噸', 'core': ['組立','鋼筋'], 'excl': ['獎金','補貼']},
    'C.03.04.00': {'unit': 'T|噸', 'core': ['鋼絲網','鋼線網'], 'excl': ['獎金']},
    'C.04.02.00': {'unit': 'M3', 'core': ['壓送'], 'excl': ['出車費','獎金']},
    'C.05.01.00': {'unit': 'M2', 'core': ['粉光','粉平'], 'excl': ['點工','獎金','角鐵','EPOXY','踢角']},
    'C.06.00.00': {'unit': 'M2', 'core': ['鷹架','鋼管','排架'], 'excl': ['母索','壁拉','樓梯','扶手','斜撐','點工','獎金','計算','扣','桁架','踏板','拉桿','防墜','腳趾','防塵','植筋','H型鋼']},
    'C.08.00.00': {'unit': 'T|噸', 'core': ['鋼構','鋼骨'], 'excl': ['防火漆','面漆','計算','鋼承板','簽證']},
    'C.08.02.00': {'unit': 'T|噸', 'core': ['鋼構','鋼骨'], 'excl': ['防火漆','面漆','計算','鋼承板']},
}

# 混凝土材料篩選 Regex
CONCRETE_MATERIAL_RE = re.compile(r'(kgf|psi|CLSM|PC混凝土|預拌混凝土|預拌SCC|膨脹混凝土|細粒料|犬走|車道植草|高壓混凝土)', re.IGNORECASE)


# ══════════════════════════════════════════
# 分類映射表
# ══════════════════════════════════════════
CATEGORY_MAP = [
    # (大項, 中項序號, 名稱, [C codes], excluded, mode)
    # mode: sum/qty/sub_header/concrete_header
    ('壹、假設工程','一','臨時水、電、話工程',['A.01.00.00'],False,'sum'),
    ('壹、假設工程','二','整地、圍籬及大門工程',['A.02.00.00','A.02.00.01'],False,'sum'),
    ('壹、假設工程','三','臨時工務所及設備',['A.03.00.00'],False,'sum'),
    ('壹、假設工程','四','勞工安全衛生費用',['A.04.00.00'],False,'sum'),
    ('壹、假設工程','五','公共設施及鄰房修護',['A.06.00.00','A.06.00.01','A.06.00.03'],False,'sum'),
    ('壹、假設工程','六','清潔及垃圾清運費',['A.05.00.00','A.05.00.01','A.05.03.00'],False,'sum'),
    ('壹、假設工程','七','營造綜合保險費',['Z.01.00.00'],False,'sum'),
    ('壹、假設工程','八','點租工',['A.08.01.00','A.08.00.00','A.08.03.00','A.08.02.00','A.08.00.01'],False,'sum'),
    ('壹、假設工程','九','營建廢棄物(含拆除/清運)',['A.05.01.00','B.06.02.00','A.08.05.00'],False,'sum'),
    ('壹、假設工程','十','假設雜項工程(含機具租金)',['A.11.00.00','A.10.00.00','N.01.00.00','A.00.00.00','A.08.06.00'],False,'sum'),

    ('貳、基礎工程','一','連續壁工程(含壁樁)',['B.04.01.00','B.04.03.00','B.02.00.00','B.02.01.00','B.02.02.00','B.02.03.00'],False,'sum'),
    ('貳、基礎工程','二','基樁工程',['B.01.00.00','B.01.02.00'],False,'sum'),
    ('貳、基礎工程','三','安全支撐工程',['B.05.00.00','B.05.01.00'],False,'sum'),
    ('貳、基礎工程','四','擋土措施',['B.03.00.00'],False,'sum'),
    ('貳、基礎工程','五','噴凝土',['B.03.07.00'],False,'sum'),
    ('貳、基礎工程','六','觀測系統工程',['B.07.00.00'],False,'sum'),
    ('貳、基礎工程','七','土方工程',['B.06.01.00','B.06.00.00'],False,'sum'),

    ('參、結構工程','一','放樣工程',['C.01.00.00','I.01.06.00'],False,'sum'),
    ('參、結構工程','二','模板',['C.02.00.00'],False,'qty'),
    ('參、結構工程','三','鋼筋',[],False,'sub_header'),
    ('參、結構工程','a','鋼筋材料',['C.03.01.00'],False,'qty'),
    ('參、結構工程','b','鋼筋加工',['C.03.02.00'],False,'qty'),
    ('參、結構工程','c','鋼筋組立',['C.03.03.00'],False,'qty'),
    ('參、結構工程','d','點焊鋼絲網工程',['C.03.04.00'],False,'qty'),
    ('參、結構工程','e','其他鋼筋工程',['C.03.05.00','C.03.00.00','C.03.00.33'],False,'sum'),
    ('參、結構工程','四','續接器工料',['C.07.00.00','C.03.07.00'],False,'sum'),
    ('參、結構工程','五','混凝土',[],False,'concrete_header'),
    ('參、結構工程','六','混凝土壓送及澆置',['C.04.02.00'],False,'qty'),
    ('參、結構工程','七','整體粉光(含混凝土標高器工料)',['C.05.01.00','C.05.02.00','C.05.00.00'],False,'qty'),
    ('參、結構工程','八','鷹架工程',['C.06.00.00'],False,'qty'),
    ('參、結構工程','九','鋼構工程',['C.08.00.00','C.08.02.00'],False,'qty'),
    ('參、結構工程','','其他結構工程',['C.00.00.00','C.10.00.00','C.09.00.00'],False,'sum'),

    ('肆、外牆裝修','一','外牆帷幕工程',['D.01.04.00'],False,'sum'),
    ('肆、外牆裝修','二','外牆泥作工程',['D.01.01.00'],False,'sum'),
    ('肆、外牆裝修','三','玻璃工程',['D.12.00.00'],False,'sum'),
    ('肆、外牆裝修','四','其他外牆裝修工程',['D.01.05.00','D.01.00.00','D.01.03.00'],False,'sum'),
    ('伍、室內裝修','一','室內泥作',['D.03.01.00'],False,'sum'),
    ('伍、室內裝修','二','隔間工程',['D.02.00.00','D.02.01.00'],False,'sum'),
    ('伍、室內裝修','三','防水',['D.08.01.00','D.08.00.00','D.08.03.00'],False,'sum'),
    ('伍、室內裝修','四','室內油漆',['D.03.05.00'],False,'sum'),
    ('伍、室內裝修','五','EPOXY',['D.03.06.00','D.03.06.02'],False,'sum'),
    ('伍、室內裝修','六','室內鐵件',['D.03.04.00'],False,'sum'),
    ('伍、室內裝修','七','其他室內裝修工程',['D.03.07.00','D.03.02.00'],False,'sum'),
    ('陸、門窗工程','一','鋁門窗工程',['D.06.00.00'],False,'sum'),
    ('陸、門窗工程','二','防火門工程',['D.07.02.00'],False,'sum'),
    ('陸、門窗工程','三','木門工程',['D.05.00.00'],False,'sum'),
    ('陸、門窗工程','四','不銹鋼門窗工程',['D.07.01.00'],False,'sum'),
    ('陸、門窗工程','五','捲門工程',['D.07.03.00'],False,'sum'),
    ('陸、門窗工程','六','其他金屬門',['D.07.04.00','D.04.00.00'],False,'sum'),
    # ✅ D.10/D.00/D.13/D.11 合併到陸門窗下
    ('陸、門窗工程','七','其他雜項工程',['D.10.00.00','D.00.00.00','D.13.00.00','D.11.00.00','D.09.00.00','D.09.02.00','D.09.05.00','D.03.03.00'],False,'sum'),

    ('柒、設備工程','一','電梯工程',['F.01.00.00'],False,'sum'),
    ('柒、設備工程','二','洗窗機工程',['F.06.00.00'],False,'sum'),
    ('柒、設備工程','三','衛浴設備',['G.05.00.00','F.03.00.00','G.00.06.00'],False,'sum'),
    ('柒、設備工程','四','其他設備工程',['F.07.00.00','F.07.00.01','F.05.00.00','F.02.00.00'],False,'sum'),
    ('捌','一','外構工程(含路面、排水)',['H.01.01.00','H.04.00.00','H.00.00.00','H.01.00.00','H.01.00.01','H.02.00.00','J.01.10.00'],False,'sum'),
    ('玖','一','景觀工程',['E.00.00.00','E.00.00.01','E.00.00.02','E.00.00.03','E.00.00.04','E.00.00.05','E.00.00.06','E.00.04.00','K.01.00.00'],False,'sum'),
    ('拾','一','機電工程',['G.00.00.00','G.01.00.00','G.06.00.00','G.02.00.00','J.04.04.00'],False,'sum'),
    ('拾壹','一','管理費',['A.10.05.00','A.10.03.00','A.10.01.00','A.10.02.00','A.10.04.00','A.10.06.00','Z.00.00.00','Z.08.00.00','Z.09.00.00','Z.00.00.01'],False,'sum'),
]

CONCRETE_CODES = ['C.04.01.00','C.04.05.00','C.04.03.00','C.04.00.00']
ALL_ASSIGNED_CODES = set()
for it in CATEGORY_MAP: ALL_ASSIGNED_CODES.update(it[3])
ALL_ASSIGNED_CODES.update(CONCRETE_CODES)

PREFIX_RULES = [
    ('A.01','壹、假設工程','一'),('A.02','壹、假設工程','二'),('A.03','壹、假設工程','三'),
    ('A.04','壹、假設工程','四'),('A.05.01','壹、假設工程','九'),('A.05','壹、假設工程','六'),
    ('A.06','壹、假設工程','五'),('A.07','壹、假設工程','七'),('A.08.05','壹、假設工程','九'),
    ('A.08.06','壹、假設工程','十'),('A.08','壹、假設工程','八'),('A.10','拾壹','一'),
    ('A.11','壹、假設工程','十'),('A.00','壹、假設工程','十'),('A.','壹、假設工程','十'),
    ('B.01','貳、基礎工程','二'),('B.02','貳、基礎工程','一'),('B.03.07','貳、基礎工程','五'),
    ('B.03','貳、基礎工程','四'),('B.04','貳、基礎工程','一'),('B.05','貳、基礎工程','三'),
    ('B.06.02','壹、假設工程','九'),('B.06','貳、基礎工程','七'),('B.07','貳、基礎工程','六'),
    ('B.','貳、基礎工程','七'),
    ('C.01','參、結構工程','一'),('C.02','參、結構工程','二'),('C.03.01','參、結構工程','a'),
    ('C.03.02','參、結構工程','b'),('C.03.03','參、結構工程','c'),('C.03.04','參、結構工程','d'),
    ('C.03','參、結構工程','e'),('C.04.02','參、結構工程','六'),('C.04','參、結構工程','五'),
    ('C.05','參、結構工程','七'),('C.06','參、結構工程','八'),('C.07','參、結構工程','四'),
    ('C.08','參、結構工程','九'),('C.','參、結構工程','五'),
    ('D.01.04','肆、外牆裝修','一'),('D.01.01','肆、外牆裝修','二'),('D.01','肆、外牆裝修','四'),
    ('D.02','伍、室內裝修','二'),('D.03.01','伍、室內裝修','一'),('D.03.04','伍、室內裝修','六'),
    ('D.03.05','伍、室內裝修','四'),('D.03.06','伍、室內裝修','五'),('D.03','伍、室內裝修','一'),
    ('D.05','陸、門窗工程','三'),('D.06','陸、門窗工程','一'),('D.07.01','陸、門窗工程','四'),
    ('D.07.02','陸、門窗工程','二'),('D.07.03','陸、門窗工程','五'),('D.07','陸、門窗工程','六'),
    ('D.08','伍、室內裝修','三'),('D.09','陸、門窗工程','七'),('D.10','陸、門窗工程','七'),
    ('D.11','陸、門窗工程','七'),('D.12','肆、外牆裝修','三'),('D.13','陸、門窗工程','七'),
    ('D.00','陸、門窗工程','七'),('D.','陸、門窗工程','七'),
    ('E.','玖','一'),('F.01','柒、設備工程','一'),('F.06','柒、設備工程','二'),
    ('F.03','柒、設備工程','三'),('G.05','柒、設備工程','三'),('G.00.06','柒、設備工程','三'),
    ('F.','柒、設備工程','四'),('G.00','拾','一'),('G.','拾','一'),
    ('H.','捌','一'),('J.01','捌','一'),('J.04','拾','一'),('J.','捌','一'),
    ('K.','玖','一'),('N.','壹、假設工程','十'),('Z.01','壹、假設工程','七'),('Z.','拾壹','一'),
]

BIG_GROUPS = [
    ('假','設',['壹、假設工程']),('基','礎',['貳、基礎工程']),('結','構',['參、結構工程']),
    ('裝','修',['肆、外牆裝修','伍、室內裝修','陸、門窗工程']),
    ('設','備',['柒、設備工程']),('景','觀',['捌','玖']),('機','電',['拾']),('管','理',['拾壹']),
]

def safe_filename(n, ml=50):
    for c in '/\\:*?"<>|': n = n.replace(c,'_')
    return n.strip()[:ml]

def find_prefix(code):
    for p,b,s in PREFIX_RULES:
        if code.startswith(p): return (b,s)
    return None

def _strip_styles(fp):
    buf = io.BytesIO()
    with zipfile.ZipFile(fp,'r') as zi:
        with zipfile.ZipFile(buf,'w',zipfile.ZIP_DEFLATED) as zo:
            for it in zi.infolist():
                d = zi.read(it.filename)
                if it.filename=='xl/styles.xml':
                    t = d.decode('utf-8',errors='replace')
                    c = max(t.count('</fill>')+t.count('<fill/>'),2)
                    s = f'<fills count="{c}"><fill><patternFill patternType="none"/></fill><fill><patternFill patternType="gray125"/></fill>'
                    for _ in range(2,c): s+='<fill><patternFill patternType="none"/></fill>'
                    s+='</fills>'
                    t = re.sub(r'<fills\b[^>]*>.*?</fills>',s,t,flags=re.DOTALL)
                    d = t.encode('utf-8')
                zo.writestr(it,d)
    buf.seek(0); return buf


def _xls_to_xlsx(filepath):
    """將 .xls 轉為 openpyxl 可讀的 BytesIO"""
    import xlrd
    xwb = xlrd.open_workbook(filepath)
    nwb = openpyxl.Workbook()
    for si, sn in enumerate(xwb.sheet_names()):
        xs = xwb.sheet_by_index(si)
        if si == 0:
            ns = nwb.active; ns.title = sn
        else:
            ns = nwb.create_sheet(title=sn)
        for r in range(xs.nrows):
            for c in range(xs.ncols):
                v = xs.cell_value(r, c)
                ct = xs.cell_type(r, c)
                if ct == xlrd.XL_CELL_EMPTY: continue
                ns.cell(r+1, c+1, v)
    buf = io.BytesIO()
    nwb.save(buf)
    buf.seek(0)
    xwb.release_resources()
    return buf


def read_kg_file(filepath):
    import warnings
    warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
    ext = os.path.splitext(filepath)[1].lower()
    if ext == '.xls':
        try:
            buf = _xls_to_xlsx(filepath)
            wb = openpyxl.load_workbook(buf, data_only=True)
        except Exception as e:
            print(f"  ⚠ .xls 讀取錯誤: {e}"); raise
    else:
        try: wb = openpyxl.load_workbook(filepath, data_only=True)
        except: wb = openpyxl.load_workbook(_strip_styles(filepath), data_only=True)
    ws = wb.active

    pname = ''
    for r in range(1,6):
        for c in range(1,10):
            v = ws.cell(r,c).value
            if v and '專案名稱' in str(v):
                nv = ws.cell(r,c+1).value
                if nv: pname = str(nv).strip()
    if not pname:
        for r in range(1,4):
            for v in [ws.cell(r,2).value, ws.cell(r,1).value]:
                if v and isinstance(v,str) and len(v.strip())>3:
                    if any(k in v for k in ['台積','根基','工程','FAB','CUP','Office']):
                        pname = v.strip(); break
            if pname: break
    if not pname: pname = os.path.splitext(os.path.basename(filepath))[0]

    # Detect headers
    gr,sr,dr,fmt = 3,4,5,'B'
    for r in range(1,15):
        rv = ' '.join(str(ws.cell(r,c).value or '') for c in range(1,45))
        if '結算完工成本' in rv or '專案任務' in rv:
            fmt = 'A' if any(k in rv for k in ['已估驗結算金額','售服費']) else 'B'
            gr,sr,dr = r,r+1,r+2; break

    hg,hs = {},{}
    for c in ws[gr]:
        if c.value: hg[c.column] = str(c.value).strip()
    for c in ws[sr]:
        if c.value: hs[c.column] = str(c.value).strip()

    col = {}
    for ci,v in hg.items():
        if v.replace(' ','')=='計價單元' and '描述' not in v: col.setdefault('C',ci)
        if '計價單元描述' in v: col.setdefault('D',ci)
        if re.search(r'工(項代碼|料編號)',v): col.setdefault('E',ci)
        if '會計科目' in v: col.setdefault('G',ci)
        if v=='單位': col.setdefault('J',ci)
        if v=='預算': col.setdefault('bud',ci)
        if '發包契約' in v: col.setdefault('ctr',ci)
        if '結算完工成本' in v: col['stl'] = ci
        if '已估驗結算金額' in v: col['Y0'] = ci
        if '未施作預估金額' in v: col['AA0'] = ci
        if '人事及規費' in v: col['AC0'] = ci
        if '售服費' in v: col['AE0'] = ci

    def fsub(s,t,mx=15):
        if s is None: return None
        for o in range(mx):
            if hs.get(s+o)==t: return s+o
        return None

    # 結算
    col['sq'] = fsub(col.get('stl'),'數量')  # settle qty
    col['sa'] = fsub(col.get('stl'),'複價')  # settle amt
    # 發包
    col['cq'] = fsub(col.get('ctr'),'數量')  # contract qty
    col['ca'] = fsub(col.get('ctr'),'複價')  # contract amt
    # 預算
    col['bq'] = fsub(col.get('bud'),'數量')
    # 四分項
    for p,k in [('Y0','Yc'),('AA0','AAc'),('AC0','ACc'),('AE0','AEc')]:
        if p in col: col[k] = fsub(col[p],'複價')

    def gc(row,key):
        ci = col.get(key)
        if ci and ci<=len(row): return row[ci-1].value
        return None
    def nm(v): return v if isinstance(v,(int,float)) else 0

    # 總計行
    total_val = 0
    tridx = ws.max_row or 5000
    for r in range(tridx,max(tridx-30,1),-1):
        dv = ws.cell(r,col.get('D',4)).value
        if dv and '總計' in str(dv):
            ci = col.get('sa')
            if ci:
                tv = ws.cell(r,ci).value
                if isinstance(tv,(int,float)): total_val = tv
            tridx = r-1; break

    # 逐列
    cat_totals = defaultdict(float)
    cat_descs = {}
    raw_rows = defaultdict(list)  # C-code → [row_data]
    conc_subs = defaultdict(lambda: {'qty':0,'amt':0,'unit':''})
    cur_c = None
    ysum=aasum=acsum=aesum=0.0

    for row in ws.iter_rows(min_row=dr, max_row=tridx, values_only=False):
        cv = gc(row,'C'); dv = gc(row,'D'); ev = gc(row,'E')
        uv = gc(row,'J'); sa = nm(gc(row,'sa')); sq = nm(gc(row,'sq'))
        cq = nm(gc(row,'cq'))

        if cv:
            cur_c = str(cv).strip()
            cur_d = str(dv).strip() if dv else ''
            if cur_c not in cat_descs and cur_d: cat_descs[cur_c] = cur_d

        ysum+=nm(gc(row,'Yc')); aasum+=nm(gc(row,'AAc'))
        acsum+=nm(gc(row,'ACc')); aesum+=nm(gc(row,'AEc'))

        if cur_c and sa:
            cat_totals[cur_c] += sa
            u = str(uv).strip() if uv else ''
            e = str(ev).strip() if ev else ''
            raw_rows[cur_c].append({'sq':sq,'sa':sa,'cq':cq,'unit':u,'code':e})

            # 混凝土子項
            if cur_c in CONCRETE_CODES and e:
                parts = e.split('-',1)
                sn = parts[1] if len(parts)>1 else e
                if CONCRETE_MATERIAL_RE.search(sn):
                    conc_subs[sn]['qty'] += sq; conc_subs[sn]['amt'] += sa
                    if u: conc_subs[sn]['unit'] = u

    vsum = sum(cat_totals.values())
    gt = (ysum+aasum+acsum+aesum) if fmt=='A' else vsum
    if total_val > 0:
        if abs(total_val - vsum) < max(vsum * 0.01, 10000):
            gt = total_val
        else:
            print(f"  ⚠ 總計行({total_val:,.0f}) ≠ 逐行加總({vsum:,.0f})，差{total_val-vsum:+,.0f}")
            print(f"    → 採用逐行加總作為結算總額")
            gt = vsum


    # ════════════════════════════════════════════
    # Format Normalizer: 空C-code行自動修正
    # ════════════════════════════════════════════
    total_data_rows = sum(1 for c,a in cat_totals.items() if a != 0)
    if total_data_rows > 0:
        strict_total = 0
        cur = None
        for row in ws.iter_rows(min_row=dr, max_row=tridx, values_only=False):
            cv = gc(row, 'C'); sa_v = nm(gc(row, 'sa'))
            if cv and str(cv).strip(): cur = str(cv).strip()
            if sa_v != 0 and bool(cv and str(cv).strip()): strict_total += sa_v
        
        inherit_total = sum(cat_totals.values())
        orphan_amt = inherit_total - strict_total
        empty_rate = orphan_amt / inherit_total * 100 if inherit_total else 0
    else:
        empty_rate = 0
    
    if empty_rate > 10:
        print(f"  ⚠ 偵測到 {empty_rate:.1f}% 金額無C-code → 啟動描述欄反查")
        
        # ── Strategy 1: 尋找高覆蓋率的「分類描述欄」(如AP6B的col29) ──
        known_descs = set(cat_descs.values())
        best_col = None; best_matches = 0
        max_col_check = min(ws.max_column or 40, 45)
        c_col_idx = col.get('C')
        
        for test_c in range(1, max_col_check + 1):
            if test_c == c_col_idx: continue
            matches = 0
            for test_r in range(dr, min(dr + 300, tridx + 1)):
                v = ws.cell(test_r, test_c).value
                if v and isinstance(v, str) and str(v).strip() in known_descs:
                    matches += 1
            if matches > best_matches:
                best_matches = matches; best_col = test_c
        
        # 檢查找到的描述欄的空行覆蓋率
        # 用該欄位自己的「值→C-code」映射（而非 known_descs）
        if best_col:
            # 從有C-code行建立 描述欄值→C-code 映射
            col_val_to_code = {}
            for row in ws.iter_rows(min_row=dr, max_row=tridx, values_only=False):
                cv = gc(row, 'C'); sa_v = nm(gc(row, 'sa'))
                if not (cv and str(cv).strip()): continue
                if sa_v == 0: continue
                fb_v = row[best_col - 1].value if best_col <= len(row) else None
                if fb_v and str(fb_v).strip():
                    val = str(fb_v).strip()
                    if val not in col_val_to_code:
                        col_val_to_code[val] = str(cv).strip()
            
            # 計算空行中有多少能被這個映射覆蓋
            empty_covered = 0; empty_total_check = 0
            for row in ws.iter_rows(min_row=dr, max_row=tridx, values_only=False):
                cv = gc(row, 'C'); sa_v = nm(gc(row, 'sa'))
                if sa_v == 0: continue
                if cv and str(cv).strip(): continue
                empty_total_check += 1
                fb_v = row[best_col - 1].value if best_col <= len(row) else None
                if fb_v and str(fb_v).strip() in col_val_to_code:
                    empty_covered += 1
            
            coverage = empty_covered / empty_total_check * 100 if empty_total_check else 0
        else:
            coverage = 0; col_val_to_code = {}
        
        # ── 覆蓋率導向搜尋（當 known_descs 搜尋覆蓋率不足時）──
        if coverage <= 80:
            # known_descs 搜尋不夠好，嘗試覆蓋率導向：
            # 對每個欄位直接建 值→C-code 映射，測試能覆蓋多少空行
            _best2 = None; _cov2 = 0; _map2 = {}
            for _tc in range(1, max_col_check + 1):
                if _tc == c_col_idx: continue
                _vm = {}; _cur = None
                for row in ws.iter_rows(min_row=dr, max_row=tridx, values_only=False):
                    cv = gc(row, 'C'); sa_v = nm(gc(row, 'sa'))
                    if cv and str(cv).strip(): _cur = str(cv).strip()
                    if not _cur or sa_v == 0: continue
                    if not (cv and str(cv).strip()): continue
                    _fb = row[_tc-1].value if _tc<=len(row) else None
                    if _fb and isinstance(_fb, str) and str(_fb).strip():
                        _vk = str(_fb).strip()
                        if _vk not in _vm: _vm[_vk] = _cur
                if not _vm: continue
                _cov_cnt = 0; _emp_cnt = 0
                for row in ws.iter_rows(min_row=dr, max_row=tridx, values_only=False):
                    cv = gc(row, 'C'); sa_v = nm(gc(row, 'sa'))
                    if sa_v == 0 or (cv and str(cv).strip()): continue
                    _emp_cnt += 1
                    _fb = row[_tc-1].value if _tc<=len(row) else None
                    if _fb and isinstance(_fb, str) and str(_fb).strip() in _vm:
                        _cov_cnt += 1
                if _emp_cnt > 0 and _cov_cnt > _cov2:
                    _cov2 = _cov_cnt; _best2 = _tc; _map2 = _vm.copy()
            _cov2_pct = _cov2 / _emp_cnt * 100 if _emp_cnt > 0 else 0
            if _cov2_pct > coverage:
                best_col = _best2; col_val_to_code = _map2
                coverage = _cov2_pct
        
        # ── 決定使用哪種策略 ──
        if coverage > 80:
            # 高覆蓋率：先檢查映射品質
            # 計算品質 = 可靠映射(>90%一對一)的比例
            _qg = 0; _qt = 0
            _val_multi = defaultdict(lambda: defaultdict(float))
            for row in ws.iter_rows(min_row=dr, max_row=tridx, values_only=False):
                cv = gc(row, 'C'); sa_v = nm(gc(row, 'sa'))
                if not (cv and str(cv).strip()) or sa_v == 0: continue
                fb_v = row[best_col-1].value if best_col<=len(row) else None
                if fb_v and str(fb_v).strip():
                    _val_multi[str(fb_v).strip()][str(cv).strip()] += abs(sa_v)
            for _vk, _codes in _val_multi.items():
                _qt += 1
                _t = sum(_codes.values()); _b = max(_codes.values())
                if _b / _t > 0.90: _qg += 1
            _map_quality = _qg / _qt * 100 if _qt else 100
            
            if _map_quality > 85:
                # ── 品質高(>85%)：直接用單一描述欄（AP6B模式，不變）──
                print(f"    找到描述欄: col {best_col} ({best_matches} matches, 覆蓋{coverage:.0f}%, 品質{_map_quality:.0f}%)")
                desc_to_code_map = col_val_to_code
                
                cat_totals_new = defaultdict(float)
                raw_rows_new = defaultdict(list)
                conc_subs_new = defaultdict(lambda: {'qty':0,'amt':0,'unit':''})
                cur_c = None
                for row in ws.iter_rows(min_row=dr, max_row=tridx, values_only=False):
                    cv = gc(row, 'C'); fb = row[best_col-1].value if best_col<=len(row) else None
                    sa_v = nm(gc(row, 'sa')); sq_v = nm(gc(row, 'sq'))
                    cq_v = nm(gc(row, 'cq')); ev = gc(row, 'E'); uv = gc(row, 'J')
                    if cv and str(cv).strip():
                        cur_c = str(cv).strip()
                    elif fb and str(fb).strip() in desc_to_code_map:
                        cur_c = desc_to_code_map[str(fb).strip()]
                    if sa_v != 0 and cur_c:
                        cat_totals_new[cur_c] += sa_v
                        u = str(uv).strip() if uv else ''; e = str(ev).strip() if ev else ''
                        raw_rows_new[cur_c].append({'sq':sq_v,'sa':sa_v,'cq':cq_v,'unit':u,'code':e})
                        if cur_c in CONCRETE_CODES and e:
                            parts = e.split('-',1); sn = parts[1] if len(parts)>1 else e
                            if CONCRETE_MATERIAL_RE.search(sn):
                                conc_subs_new[sn]['qty']+=sq_v; conc_subs_new[sn]['amt']+=sa_v
                                if u: conc_subs_new[sn]['unit']=u
                cat_totals = dict(cat_totals_new); raw_rows = dict(raw_rows_new); conc_subs = dict(conc_subs_new)
                print(f"    修正後 C-codes: {len(cat_totals)}")
            else:
                # ── 品質低(≤85%)：多欄位品質導向（零廢中心模式）──
                print(f"    描述欄 col {best_col} 覆蓋{coverage:.0f}% 但品質{_map_quality:.0f}% → 多欄位品質導向")
                # 對所有欄位建立 值→(C-code, 品質%) 映射
                _all_maps = {}
                for _tc in range(1, max_col_check + 1):
                    if _tc == c_col_idx: continue
                    _vc = defaultdict(lambda: defaultdict(float))
                    for row in ws.iter_rows(min_row=dr, max_row=tridx, values_only=False):
                        cv = gc(row, 'C'); sa_v = nm(gc(row, 'sa'))
                        if not (cv and str(cv).strip()) or sa_v == 0: continue
                        _fb = row[_tc-1].value if _tc<=len(row) else None
                        if _fb and isinstance(_fb, str) and str(_fb).strip():
                            _vc[str(_fb).strip()][str(cv).strip()] += abs(sa_v)
                    if not _vc: continue
                    _vm = {}
                    for _vk, _codes in _vc.items():
                        _t = sum(_codes.values()); _b = max(_codes.items(), key=lambda x:x[1])
                        _vm[_vk] = (_b[0], _b[1]/_t*100)
                    if _vm: _all_maps[_tc] = _vm
                
                cat_totals_new = defaultdict(float)
                raw_rows_new = defaultdict(list)
                conc_subs_new = defaultdict(lambda: {'qty':0,'amt':0,'unit':''})
                cur_c = None
                for row in ws.iter_rows(min_row=dr, max_row=tridx, values_only=False):
                    cv = gc(row, 'C')
                    sa_v = nm(gc(row, 'sa')); sq_v = nm(gc(row, 'sq'))
                    cq_v = nm(gc(row, 'cq')); ev = gc(row, 'E'); uv = gc(row, 'J')
                    if cv and str(cv).strip():
                        cur_c = str(cv).strip()
                    else:
                        _bc = None; _bq = 0
                        for _tc, _vm in _all_maps.items():
                            _fb = row[_tc-1].value if _tc<=len(row) else None
                            if _fb and isinstance(_fb, str) and str(_fb).strip() in _vm:
                                _code, _q = _vm[str(_fb).strip()]
                                if _q > _bq: _bq = _q; _bc = _code
                        if _bc and _bq > 70: cur_c = _bc
                    if sa_v != 0 and cur_c:
                        cat_totals_new[cur_c] += sa_v
                        u = str(uv).strip() if uv else ''; e = str(ev).strip() if ev else ''
                        raw_rows_new[cur_c].append({'sq':sq_v,'sa':sa_v,'cq':cq_v,'unit':u,'code':e})
                        if cur_c in CONCRETE_CODES and e:
                            parts = e.split('-',1); sn = parts[1] if len(parts)>1 else e
                            if CONCRETE_MATERIAL_RE.search(sn):
                                conc_subs_new[sn]['qty']+=sq_v; conc_subs_new[sn]['amt']+=sa_v
                                if u: conc_subs_new[sn]['unit']=u
                cat_totals = dict(cat_totals_new); raw_rows = dict(raw_rows_new); conc_subs = dict(conc_subs_new)
                print(f"    修正後 C-codes: {len(cat_totals)}")
        else:
            # 低覆蓋率：三層 fallback（零廢中心模式）
            print(f"    描述欄覆蓋率低({coverage:.0f}%) → 啟用三層fallback")
            
            # Layer 1: 描述→C-code（從有C-code行建立）
            desc_to_code_map = {}
            for c_code, desc in cat_descs.items():
                if desc not in desc_to_code_map: desc_to_code_map[desc] = c_code
            
            # Layer 2: 合約→C-code（從有C-code行建立）
            contract_codes_map = defaultdict(lambda: defaultdict(float))
            cur_code = None
            for row in ws.iter_rows(min_row=dr, max_row=tridx, values_only=False):
                cv = gc(row, 'C'); sa_v = nm(gc(row, 'sa'))
                # 動態找合約欄
                ctr_ci = col.get('ctr')
                if ctr_ci:
                    for off in range(6):
                        sv = ws.cell(dr-1, ctr_ci+off).value  # sub header row
                        if sv and '合約編號' in str(sv): ctr_ci = ctr_ci+off; break
                ctr_v = row[ctr_ci-1].value if ctr_ci and ctr_ci<=len(row) else None
                
                if cv and str(cv).strip(): cur_code = str(cv).strip()
                if not isinstance(sa_v,(int,float)) or sa_v==0: continue
                if not cur_code: continue
                if cv and str(cv).strip() and ctr_v:
                    contract_codes_map[str(ctr_v).strip()][cur_code] += abs(sa_v)
            
            contract_best_map = {}
            for ctr, codes in contract_codes_map.items():
                best = max(codes.items(), key=lambda x: x[1])
                contract_best_map[ctr] = best[0]
            
            # 重建
            cat_totals_new = defaultdict(float)
            raw_rows_new = defaultdict(list)
            conc_subs_new = defaultdict(lambda: {'qty':0,'amt':0,'unit':''})
            cur_c = None
            d_col = col.get('D')  # 計價單元描述欄
            fix_d = fix_ctr = fix_inh = 0
            
            for row in ws.iter_rows(min_row=dr, max_row=tridx, values_only=False):
                cv = gc(row, 'C'); sa_v = nm(gc(row, 'sa'))
                sq_v = nm(gc(row, 'sq')); cq_v = nm(gc(row, 'cq'))
                ev = gc(row, 'E'); uv = gc(row, 'J')
                dv = gc(row, 'D')  # 計價單元描述
                ctr_ci = col.get('ctr')
                if ctr_ci:
                    for off in range(6):
                        sv = ws.cell(dr-1, ctr_ci+off).value
                        if sv and '合約編號' in str(sv): ctr_ci = ctr_ci+off; break
                ctr_v = row[ctr_ci-1].value if ctr_ci and ctr_ci<=len(row) else None
                
                if cv and str(cv).strip():
                    cur_c = str(cv).strip()
                else:
                    d = str(dv).strip() if dv else ''
                    c_str = str(ctr_v).strip() if ctr_v else ''
                    if d and d in desc_to_code_map:
                        cur_c = desc_to_code_map[d]; fix_d += 1
                    elif c_str and c_str in contract_best_map:
                        cur_c = contract_best_map[c_str]; fix_ctr += 1
                    else:
                        fix_inh += 1
                
                if sa_v != 0 and cur_c:
                    cat_totals_new[cur_c] += sa_v
                    u = str(uv).strip() if uv else ''; e = str(ev).strip() if ev else ''
                    raw_rows_new[cur_c].append({'sq':sq_v,'sa':sa_v,'cq':cq_v,'unit':u,'code':e})
                    if cur_c in CONCRETE_CODES and e:
                        parts = e.split('-',1); sn = parts[1] if len(parts)>1 else e
                        if CONCRETE_MATERIAL_RE.search(sn):
                            conc_subs_new[sn]['qty']+=sq_v; conc_subs_new[sn]['amt']+=sa_v
                            if u: conc_subs_new[sn]['unit']=u
            
            cat_totals = dict(cat_totals_new); raw_rows = dict(raw_rows_new); conc_subs = dict(conc_subs_new)
            print(f"    描述反查={fix_d} 合約反查={fix_ctr} 繼承={fix_inh}")
            print(f"    修正後 C-codes: {len(cat_totals)}")
    wb.close()
    return {
        'project_name':pname,'cat_totals':dict(cat_totals),'cat_descs':cat_descs,
        'raw_rows':dict(raw_rows),'concrete_subs':dict(conc_subs),
        'settlement':{'fmt':fmt,'total':gt,'V':vsum},
    }


def extract_qty(raw_rows_for_code, code):
    """用核心工項篩選提取代表性數量"""
    rules = QTY_RULES.get(code)
    if not rules: return 0, ''
    
    target_units = rules['unit'].split('|')
    core_kw = rules['core']
    excl_kw = rules['excl']
    
    total_cq = 0  # 發包數量
    total_sq = 0  # 結算數量
    best_unit = ''
    
    for rd in raw_rows_for_code:
        u = rd['unit']
        desc = rd['code']
        
        # 單位必須匹配
        if not any(u.upper().startswith(tu.upper()) for tu in target_units):
            continue
        # 排除非核心
        if any(kw in desc for kw in excl_kw):
            continue
        
        cq = rd['cq']
        sq = rd['sq']
        if cq > 0: total_cq += cq
        if sq > 0: total_sq += sq
        if u and not best_unit: best_unit = u
    
    # 優先發包數量，否則結算數量
    qty = total_cq if total_cq > 0 else total_sq
    return qty, best_unit


def classify_project(proj):
    ct = proj['cat_totals']; ds = proj['cat_descs']
    rr = proj['raw_rows']; cs = proj['concrete_subs']

    cidx = {}
    for i,it in enumerate(CATEGORY_MAP): cidx[(it[0],it[1])] = i

    res = []
    for b,s,n,codes,ex,m in CATEGORY_MAP:
        res.append({'big':b,'sub':s,'name':n,'codes':list(codes),
                    'total':0,'excluded':ex,'mode':m,
                    'qty':0,'unit':'','fb':[],'conc_items':[]})

    assigned = set(ALL_ASSIGNED_CODES); flog = []

    # 精確
    for code,amt in ct.items():
        if code in assigned:
            for i,(b,s,n,codes,ex,m) in enumerate(CATEGORY_MAP):
                if code in codes:
                    res[i]['total'] += amt
                    # qty提取
                    if m=='qty' and code in rr:
                        q,u = extract_qty(rr[code],code)
                        res[i]['qty'] += q
                        if u and not res[i]['unit']: res[i]['unit'] = u
                    break
            if code in CONCRETE_CODES:
                for i,it in enumerate(CATEGORY_MAP):
                    if it[5]=='concrete_header':
                        res[i]['total'] += amt; break

    # 前綴
    for code,amt in ct.items():
        if code not in assigned and amt!=0:
            t = find_prefix(code)
            if t and t in cidx:
                idx = cidx[t]; res[idx]['total'] += amt
                res[idx]['fb'].append(code); assigned.add(code)
                flog.append(f"  {code} ({ds.get(code,'')}) → {t}")

    unassigned = {c:{'amount':a,'desc':ds.get(c,'')} for c,a in ct.items() if c not in assigned and a!=0}

    # 混凝土子項
    for i,it in enumerate(CATEGORY_MAP):
        if it[5]=='concrete_header':
            items = []
            # 排序
            order_kw = ['犬走','車道','高壓','CLSM','420','560kgf/c㎡預拌混凝土','560kgf/c㎡預拌SCC',
                        '細粒料','140','210','280kgf/c㎡預拌混凝土','280kgf/c㎡預拌膨脹','280kgf/c㎡SCC']
            used = set()
            for kw in order_kw:
                for sn,sv in cs.items():
                    if kw in sn and sn not in used:
                        items.append((sn,sv)); used.add(sn)
            for sn,sv in cs.items():
                if sn not in used: items.append((sn,sv))
            # 未匹配 Regex 的金額歸入「其他」
            matched_amt = sum(sv['amt'] for _,sv in items)
            total_conc = sum(ct.get(c,0) for c in CONCRETE_CODES)
            other_amt = total_conc - matched_amt
            if other_amt > 0:
                items.append(('其他混凝土工程',{'qty':0,'amt':other_amt,'unit':'式'}))
            res[i]['conc_items'] = items
            break

    # 合計（只排除 sub_header，concrete_header 的金額要計入）
    skip_modes = {'sub_header'}
    gt = sum(r['total'] for r in res if r['mode'] not in skip_modes)
    gi = sum(r['total'] for r in res if not r['excluded'] and r['mode'] not in skip_modes)

    return {'items':res,'gt':gt,'gi':gi,'unassigned':unassigned,'flog':flog}


def write_cost_analysis(wb, proj, clf, area_m2, sname):
    ws = wb.create_sheet(title=sname)
    ap = area_m2*PING_PER_M2 if area_m2 else 0

    ft = Font(name='微軟正黑體',bold=True,size=14)
    fc = Font(name='微軟正黑體',size=12)
    fp = Font(name='微軟正黑體',size=12)
    fb = Font(name='微軟正黑體',size=10)
    fi = Font(name='微軟正黑體',size=8.5)
    fv = Font(name='微軟正黑體',size=10)
    ftl = Font(name='微軟正黑體',bold=True,size=10)
    fm = Font(name='微軟正黑體',size=8,color='CC6600')
    fs = Font(name='微軟正黑體',size=10)
    fh = Font(name='微軟正黑體',size=8)

    thin = Side(style='thin'); med = Side(style='medium')
    bt = Border(top=thin,bottom=thin,left=thin,right=thin)

    for c,w in {'A':14.8,'B':2.8,'C':10.4,'D':3.6,'E':8.2,'F':11.8,'G':3.0,
                'H':2.8,'I':12.0,'J':2.8,'K':2.8,'L':10.9,'M':5.1,'N':3.8,
                'O':11.8,'P':3.6,'Q':2.8,'R':9.9,'S':3.9,'T':3.0,'U':9.9,
                'V':1.8,'W':3.0,'X':16.6}.items():
        ws.column_dimensions[c].width = w

    # 工程造價合計 = 結算總額（不排除任何項目）
    settle_total = proj['settlement']['total']
    gi = settle_total if settle_total > 0 else clf['gi']

    # Headers
    ws['A1']='工 程 造 價 統 計 ( 分 析 ) 表'; ws['A1'].font=ft
    ws.merge_cells('A2:F2'); ws['A2']='根基營造股份有限公司'; ws['A2'].font=fc
    ws['W2']=datetime.now().strftime('%Y年 %m月%d日'); ws['W2'].font=fs
    ws.merge_cells('A3:D3'); ws['A3']='  工 程 名 稱'; ws['A3'].font=fp; ws.merge_cells('E3:I3')
    ws['E3']=proj['project_name']; ws['E3'].font=fp
    ws.merge_cells('J3:L3'); ws['J3']='樓  層  面  積'; ws['J3'].font=fs
    if area_m2: ws['M3']=f'{area_m2:,.2f} m²'; ws['M3'].font=fs
    ws['X3']=ap; ws['X3'].number_format='#,##0.00'; ws['X3'].font=fs
    ws.merge_cells('A4:A5'); ws['A4']='概要'; ws['A4'].font=fb

    # Row 6-7
    bgt = []
    for _,_,bl in BIG_GROUPS:
        # 包含所有項目（含concrete_header的混凝土合計），只排除sub_header（金額=0的標題行）
        t = sum(it['total'] for it in clf['items'] if it['big'] in bl and it['mode'] != 'sub_header')
        bgt.append(t)
    ws.merge_cells('A6:A7'); ws['A6']='大項'; ws['A6'].font=fs
    scc = [2,5,8,11,14,17,20,23]
    svm = [('C','D'),('F','G'),('I','J'),('L','M'),('O','P'),('R','S'),('U','V'),('X','Y')]
    for i,(c1,c2,_) in enumerate(BIG_GROUPS):
        t=bgt[i]; pp=round(t/ap) if ap else 0; pct=t/gi if gi else 0
        cc=scc[i]; v1=openpyxl.utils.column_index_from_string(svm[i][0]); v2=openpyxl.utils.column_index_from_string(svm[i][1])
        ws.cell(6,cc,c1).font=fs; ws.merge_cells(start_row=6,start_column=v1,end_row=6,end_column=v2)
        ws.cell(6,v1,f'{pp:,}元／坪').font=fs
        ws.cell(7,cc,c2).font=fs; ws.merge_cells(start_row=7,start_column=v1,end_row=7,end_column=v2)
        ws.cell(7,v1).value=pct; ws.cell(7,v1).number_format='0.00%'; ws.cell(7,v1).font=fs

    # Row 8-10
    ws.merge_cells('A8:N8'); ws['A8']='分   類   造   價   及   數   量   統   計   表           '; ws['A8'].font=ftl
    for c in range(1,25): ws.cell(8,c).border=Border(top=med)
    ws.merge_cells('R8:S8'); ws['R8']='工程造價:'; ws['R8'].font=ftl
    if ap: ws['T8']=gi/ap; ws['T8'].number_format='#,##0.00'; ws['T8'].font=fs
    ws['W8']='元／坪(未稅)'; ws['W8'].font=fs
    ws.merge_cells('A9:A10'); ws['A9']='大項'; ws['A9'].font=fs; ws.merge_cells('B9:B10'); ws['B9']='中項'; ws['B9'].font=fs
    ws.cell(10,3,'項       目').font=fs; ws.cell(10,18,' 權   比').font=fs
    for c in range(1,20): ws.cell(10,c).border=Border(bottom=med)

    def wr(r,big,sub,name,qty,unit,amt,mode='sum',hdr=False,font=None):
        fn=font or fi
        if big: ws.cell(r,1,big).font=fb
        ws.cell(r,2,sub).font=fv
        ws.merge_cells(start_row=r,start_column=3,end_row=r,end_column=5)
        ws.cell(r,3,name).font=fn
        if hdr:
            for c in range(1,20): ws.cell(r,c).border=bt
            return
        ws.cell(r,6,qty).font=fv
        if isinstance(qty,(int,float)) and qty!=0: ws.cell(r,6).number_format='#,##0.00'
        ws.merge_cells(start_row=r,start_column=7,end_row=r,end_column=8)
        ws.cell(r,7,unit).font=fv
        if mode=='qty' and isinstance(qty,(int,float)) and qty>0:
            up = amt/qty
            ws.cell(r,9,round(up,1)).font=fv
        else:
            ws.cell(r,9,amt).font=fv
        ws.cell(r,9).number_format='#,##0.0'
        ws.merge_cells(start_row=r,start_column=10,end_row=r,end_column=11)
        ws.cell(r,10,'元').font=fv
        pm2=amt/area_m2 if area_m2 else 0; pp=amt/ap if ap else 0; pct=amt/gi if gi else 0
        ws.merge_cells(start_row=r,start_column=13,end_row=r,end_column=14)
        ws.merge_cells(start_row=r,start_column=16,end_row=r,end_column=17)
        ws.merge_cells(start_row=r,start_column=19,end_row=r,end_column=25)
        ws.cell(r,12,pm2).font=fv; ws.cell(r,12).number_format='#,##0.00'
        ws.cell(r,13,'元/m2').font=fv
        ws.cell(r,15,pp).font=fv; ws.cell(r,15).number_format='#,##0.00'
        ws.cell(r,16,'元/坪').font=fv
        ws.cell(r,18,pct).font=fv; ws.cell(r,18).number_format='0.00%'
        for c in range(1,20): ws.cell(r,c).border=bt

    row=11; cb=None; bs=None; bsecs=[]

    for it in clf['items']:
        m=it['mode']
        if m=='sub_header': continue
        if m=='concrete_header':
            if not it['conc_items'] and it['total']==0: continue
            nf=(it['big']!=cb)
            if nf:
                if cb and bs: bsecs.append((bs,row-1,cb))
                cb=it['big']; bs=row
            wr(row,it['big'] if nf else '',it['sub'],it['name'],'','',0,hdr=True); row+=1
            alpha='abcdefghijklmnopqrstuvwxyz'
            for si,(sn,sd) in enumerate(it['conc_items']):
                sl=alpha[si] if si<26 else str(si)
                sq=round(sd['qty'],1) if sd['qty'] else 1
                su=sd['unit'] or 'm3'
                sa=round(sd['amt'])
                if sa==0: continue
                if sn=='其他混凝土工程': wr(row,'',sl,sn,1,'式',sa,mode='sum')
                else: wr(row,'',sl,sn,sq,su,sa,mode='qty')
                row+=1
            continue

        if it['total']==0 and not it['fb']: continue
        nf=(it['big']!=cb)
        if nf:
            if cb and bs: bsecs.append((bs,row-1,cb))
            cb=it['big']; bs=row

        # 鋼筋 header
        if it['sub']=='a' and it['big']=='參、結構工程':
            wr(row,it['big'] if nf else '','三','鋼筋','','',0,hdr=True); nf=False; row+=1

        name=it['name']
        if it['fb']: name+=f" (+{','.join(it['fb'])})"

        if m=='qty' and it['qty']>0:
            wr(row,it['big'] if nf else '',it['sub'],name,round(it['qty'],2),it['unit'],round(it['total']),mode='qty')
        else:
            wr(row,it['big'] if nf else '',it['sub'],name,1,'式',round(it['total']),mode='sum')
        row+=1

    # 未歸類
    ua=clf.get('unassigned',{})
    if ua:
        if cb and bs: bsecs.append((bs,row-1,cb))
        cb='雜項工程(未歸類)'; bs=row
        ws.cell(row,1,'雜項工程(未歸類)').font=Font(name='微軟正黑體',size=10,color='CC6600')
        for code,info in sorted(ua.items(),key=lambda x:-abs(x[1]['amount'])):
            wr(row,'','',f"{code} {info['desc']}",1,'式',round(info['amount']),font=fm)
            row+=1

    if cb and bs: bsecs.append((bs,row-1,cb))
    for s,e,_ in bsecs:
        if e>s: ws.merge_cells(start_row=s,start_column=1,end_row=e,end_column=1)

    # 合計
    for mc in [(3,5),(7,8),(10,11),(13,14),(16,17),(19,25)]:
        ws.merge_cells(start_row=row,start_column=mc[0],end_row=row,end_column=mc[1])
    ws.cell(row,1,'合         計').font=ftl
    ws.cell(row,9,gi).font=ftl; ws.cell(row,9).number_format='#,##0'
    tpm2=gi/area_m2 if area_m2 else 0; tpp=gi/ap if ap else 0
    ws.cell(row,12,tpm2).font=ftl; ws.cell(row,12).number_format='#,##0.00'
    ws.cell(row,13,'元/m2').font=fv; ws.cell(row,15,tpp).font=ftl; ws.cell(row,15).number_format='#,##0.00'
    ws.cell(row,16,'元/坪').font=fv; ws.cell(row,18,1).font=ftl; ws.cell(row,18).number_format='0.00%'
    ws.cell(row,19,f'{gi:,.0f}未稅   (結算)').font=fv
    for c in range(1,20): ws.cell(row,c).border=Border(top=thin,bottom=med,left=thin,right=thin)
    for c in range(1,20):
        cl=ws.cell(11,c); o=cl.border
        cl.border=Border(top=med,bottom=o.bottom,left=o.left,right=o.right)


def ext_name(fp):
    fn=os.path.splitext(os.path.basename(fp))[0]
    n=re.sub(r'^KG預算發包請款?結算一覽表[_\s]*','',fn)
    n=re.sub(r'^KG[_\s]*','',n)
    n=re.sub(r'[_\s]+$','',re.sub(r'^[_\s]+','',n)) or fn
    # 移除括號，替換空格為底線
    n=n.replace('(','').replace(')','').replace(' ','_')
    return n


def main():
    pa=argparse.ArgumentParser(description='KG→造價分析 v12')
    pa.add_argument('--input_dir','-i',default='.')
    pa.add_argument('--output_dir','-O',default=None)
    pa.add_argument('--area_file','-a',default=None)
    pa.add_argument('--default_area','-d',type=float,default=None)
    args=pa.parse_args()

    kf = []
    for ext in ['*.xlsx', '*.xls']:
        kf += glob.glob(os.path.join(args.input_dir, f'KG預算發包請款結算一覽表{ext}'))
        kf += glob.glob(os.path.join(args.input_dir, f'KG{ext}'))
    kf = sorted(set(kf))
    if not kf: print(f"找不到 KG 檔案 (.xlsx/.xls): {args.input_dir}"); sys.exit(1)
    print(f"找到 {len(kf)} 個 KG 檔案")

    am={}
    if args.area_file and os.path.exists(args.area_file):
        with open(args.area_file,'r',encoding='utf-8-sig') as f:
            rd=csv.reader(f); next(rd,None)
            for r in rd:
                if len(r)>=2:
                    try: am[r[0].strip()]=float(r[1].strip())
                    except: pass

    for idx,fp in enumerate(kf):
        fn=os.path.basename(fp); print(f"\n[{idx+1}/{len(kf)}] {fn}")
        try: proj=read_kg_file(fp)
        except Exception as e: print(f"  錯誤: {e}"); continue
        sb=proj['settlement']
        print(f"  專案: {proj['project_name']} | C codes: {len(proj['cat_totals'])} | 結算: {sb['total']:,.0f}")

        a=am.get(fn,args.default_area)
        if a is None:
            try:
                inp=input(f"  請輸入面積m²: ").strip()
                a=float(inp) if inp else 0
            except: a=0

        clf=classify_project(proj)
        t=sb['total'] if sb['total']>0 else clf['gi']
        ap=a*PING_PER_M2 if a else 0
        if a: print(f"  {t:,.0f} | {a:,.0f}m² | {t/ap:,.0f}元/坪")
        if clf['unassigned']: print(f"  ⚠ 未歸類 {len(clf['unassigned'])}項")

        pl=ext_name(fp); sf=safe_filename(pl) or f'P{idx+1}'
        td=datetime.now().strftime('%Y%m%d')

        wb=openpyxl.Workbook(); wb.remove(wb.active)
        write_cost_analysis(wb,proj,clf,a,f'{sf}_造價'[:31])

        ofn=f'{sf}_{td}.xlsx'
        if args.output_dir: os.makedirs(args.output_dir,exist_ok=True); ofn=os.path.join(args.output_dir,ofn)
        # 儲存（若檔案被鎖定則自動加序號）
        saved = False
        for attempt in range(10):
            try:
                save_path = ofn if attempt == 0 else ofn.replace('.xlsx', f'_{attempt}.xlsx')
                wb.save(save_path)
                print(f"  → {save_path}")
                saved = True
                break
            except PermissionError:
                if attempt == 0:
                    print(f"  ⚠ {ofn} 被鎖定（可能已在Excel中開啟）")
        if not saved:
            print(f"  ❌ 無法儲存，請關閉Excel後重試")

    print(f"\n完成！")

if __name__=='__main__': main()
